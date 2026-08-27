from __future__ import annotations

import io
import logging

from openpyxl import load_workbook

from app.services.files._common import parse_in_thread

logger = logging.getLogger(__name__)

_MAX_CHARS = 60_000


def _extract_text(buffer: bytes) -> str:
    # Открываем книгу дважды: с data_only=True — для значений, и с
    # data_only=False — только чтобы отличить формулу без закэшированного
    # значения от реально пустой ячейки (при data_only=True обе выглядят как
    # value=None, data_type="n" — неотличимо). Обе книги в read_only=True,
    # поэтому чтение потоковое и раннее прерывание ниже реально экономит I/O,
    # а не только Python-обработку.
    workbook = load_workbook(io.BytesIO(buffer), read_only=True, data_only=True)
    try:
        formula_workbook = load_workbook(io.BytesIO(buffer), read_only=True, data_only=False)
        try:
            parts: list[str] = []
            total_len = 0
            lost_formula_count = 0

            for sheet, formula_sheet in zip(workbook.worksheets, formula_workbook.worksheets):
                rows = sheet.iter_rows(values_only=True)
                formula_rows = formula_sheet.iter_rows(values_only=False)
                for row, formula_row in zip(rows, formula_rows):
                    if all(cell is None for cell in row):
                        continue
                    cells: list[str] = []
                    for value, formula_cell in zip(row, formula_row):
                        if value is None and formula_cell.data_type == "f":
                            # Формула без закэшированного значения (файл
                            # собран скриптом и никогда не открывался в
                            # Excel/LibreOffice для пересчёта) — данные
                            # реально потеряны, а не просто пусты.
                            lost_formula_count += 1
                            cells.append("")
                        else:
                            cells.append("" if value is None else str(value))
                    line = "\t".join(cells)
                    parts.append(line)
                    total_len += len(line) + 1
                    if total_len >= _MAX_CHARS:
                        break
                if total_len >= _MAX_CHARS:
                    break

            if lost_formula_count:
                logger.warning(
                    "xlsx contains %d formula cell(s) without a cached value (buffer %d bytes) — "
                    "their content could not be extracted",
                    lost_formula_count,
                    len(buffer),
                )

            # strip("\n"), не strip() без аргументов: последняя ячейка
            # последней строки может законно быть пустой (например, "потерянная"
            # формула выше) — bare strip() съел бы завершающий "\t" и сдвинул
            # бы колонки. Обрезаем только случайные пустые строки по краям.
            return "\n".join(parts).strip("\n")[:_MAX_CHARS]
        finally:
            formula_workbook.close()
    finally:
        workbook.close()


async def parse_xlsx(buffer: bytes) -> str:
    """Извлекает обычный текст из XLSX-буфера. Никогда не бросает исключения:
    любая ошибка (битая книга, не-xlsx байты, ошибка парсера) приводит
    к пустой строке, чтобы вызывающий код мог безопасно откатиться на заглушку.

    Каждая непустая строка таблицы становится одной строкой результата,
    ячейки внутри неё склеены табуляцией; строки, где все ячейки пустые
    (None), полностью выбрасываются из результата.

    Если в книге есть формулы без закэшированного значения (файл ни разу не
    открывался в Excel/LibreOffice для пересчёта), такие ячейки читаются как
    пустые и это логируется через `logger.warning` — восстановить значение
    формулы средствами openpyxl невозможно, это не ошибка парсинга.

    Разбор синхронный и CPU-bound (у openpyxl нет async API), поэтому он
    выполняется в отдельном потоке через `asyncio.to_thread`, чтобы не
    блокировать event loop на время разбора большого файла.
    """
    return await parse_in_thread(_extract_text, buffer, logger, "parse_xlsx")
