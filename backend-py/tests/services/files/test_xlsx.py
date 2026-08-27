import io
import logging

from openpyxl import Workbook

from app.services.files.xlsx import parse_xlsx

GARBAGE_BUFFER = b"this is definitely not an xlsx file at all"


def _xlsx_with_rows_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Name", "Age"])
    sheet.append(["Alice", 30])
    sheet.append([None, None])
    sheet.append(["Bob", 25])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


async def test_parse_xlsx_returns_tab_separated_rows_and_skips_empty_rows():
    result = await parse_xlsx(_xlsx_with_rows_bytes())
    lines = result.split("\n")
    assert "Name\tAge" in lines
    assert "Alice\t30" in lines
    assert "Bob\t25" in lines
    # Строка, где все ячейки пустые (None), не должна попасть в результат
    # ни как содержательная строка, ни как пустая строка между другими.
    assert "" not in lines
    assert len(lines) == 3


async def test_parse_xlsx_never_throws_on_non_xlsx_buffer():
    result = await parse_xlsx(GARBAGE_BUFFER)
    assert result == ""


async def test_parse_xlsx_never_throws_on_empty_buffer():
    result = await parse_xlsx(b"")
    assert result == ""


async def test_parse_xlsx_warns_and_blanks_uncached_formula(caplog):
    # Регрессия: load_workbook(..., data_only=True) отдаёт value=None для
    # формулы, которую никогда не пересчитывал Excel/LibreOffice — это
    # неотличимо от реально пустой ячейки без дополнительной проверки.
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Name", "Total"])
    sheet.append(["Alice", "=1+1"])
    buffer = io.BytesIO()
    workbook.save(buffer)

    with caplog.at_level(logging.WARNING, logger="app.services.files.xlsx"):
        result = await parse_xlsx(buffer.getvalue())

    lines = result.split("\n")
    assert "Name\tTotal" in lines
    # Ячейка формулы стала пустой строкой, но строка не выброшена целиком.
    assert "Alice\t" in lines
    assert any("formula cell" in record.message for record in caplog.records)


async def test_parse_xlsx_truncates_incrementally(monkeypatch):
    monkeypatch.setattr("app.services.files.xlsx._MAX_CHARS", 20)
    workbook = Workbook()
    sheet = workbook.active
    for i in range(50):
        sheet.append([f"row{i}", "some longer text value"])
    buffer = io.BytesIO()
    workbook.save(buffer)

    result = await parse_xlsx(buffer.getvalue())
    assert len(result) <= 20
